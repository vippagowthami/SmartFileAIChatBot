import os
import re
import json
import csv
from io import StringIO
from pathlib import Path


class FileProcessor:
    """Handles file upload and text extraction for multiple formats"""

    SUPPORTED_FORMATS = {".txt", ".pdf", ".doc", ".docx", ".csv", ".json", ".xlsx", ".xls"}
    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB
    # Approximate chars per token for English (1 token ≈ 4 chars)
    CHARS_PER_TOKEN = 4

    @staticmethod
    def extract_text(file_path: str) -> str:
        file_ext = Path(file_path).suffix.lower()
        if file_ext == ".txt":
            return FileProcessor._extract_txt(file_path)
        elif file_ext == ".pdf":
            return FileProcessor._extract_pdf(file_path)
        elif file_ext in {".doc", ".docx"}:
            return FileProcessor._extract_docx(file_path)
        elif file_ext == ".csv":
            return FileProcessor._extract_csv(file_path)
        elif file_ext == ".json":
            return FileProcessor._extract_json(file_path)
        elif file_ext in {".xlsx", ".xls"}:
            return FileProcessor._extract_excel(file_path)
        else:
            raise ValueError(f"Unsupported file format: {file_ext}")

    @staticmethod
    def _extract_txt(file_path: str) -> str:
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                return f.read()
        except UnicodeDecodeError:
            with open(file_path, "r", encoding="latin-1") as f:
                return f.read()

    @staticmethod
    def _extract_pdf(file_path: str) -> str:
        try:
            import PyPDF2
            text = []
            with open(file_path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text.append(page_text)
            return "\n\n".join(text)
        except ImportError:
            raise ImportError("PyPDF2 required. Run: pip install PyPDF2")
        except Exception as e:
            raise Exception(f"Failed to extract PDF: {e}")

    @staticmethod
    def _extract_docx(file_path: str) -> str:
        try:
            from docx import Document
            doc = Document(file_path)
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n\n".join(paragraphs)
        except ImportError:
            raise ImportError("python-docx required. Run: pip install python-docx")
        except Exception as e:
            raise Exception(f"Failed to extract DOCX: {e}")

    @staticmethod
    def _extract_csv(file_path: str) -> str:
        """Extract CSV data into readable format with headers and rows."""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                if reader.fieldnames is None:
                    return ""
                
                rows = list(reader)
                if not rows:
                    return ""
                
                # Build readable format: headers and rows
                lines = [f"CSV Data: {Path(file_path).name}"]
                lines.append(f"Headers: {', '.join(reader.fieldnames)}")
                lines.append("")
                
                for idx, row in enumerate(rows[:1000], 1):  # Limit to 1000 rows
                    row_str = " | ".join(f"{k}: {v}" for k, v in row.items())
                    lines.append(f"Row {idx}: {row_str}")
                
                return "\n".join(lines)
        except Exception as e:
            raise Exception(f"Failed to extract CSV: {e}")

    @staticmethod
    def _extract_json(file_path: str) -> str:
        """Extract JSON data into readable format."""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            
            # Convert to readable text
            if isinstance(data, dict):
                return json.dumps(data, indent=2)
            elif isinstance(data, list):
                return "\n".join(json.dumps(item, indent=2) for item in data)
            else:
                return str(data)
        except Exception as e:
            raise Exception(f"Failed to extract JSON: {e}")

    @staticmethod
    def _extract_excel(file_path: str) -> str:
        """Extract Excel data into readable format."""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(file_path, data_only=True)
            all_text = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                all_text.append(f"\n[Sheet: {sheet_name}]")
                
                # Extract all rows
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
                        all_text.append(row_str)
            
            return "\n".join(all_text)
        except ImportError:
            raise ImportError("openpyxl required. Run: pip install openpyxl")
        except Exception as e:
            raise Exception(f"Failed to extract Excel: {e}")

    @staticmethod
    def chunk_text(text: str, chunk_size: int = 600, overlap: int = 120) -> list[str]:
        """
        Smart chunking optimized for LOCAL LLM (400-700 tokens, 100-150 overlap).
        
        Strategy:
        - chunk_size: ~600 tokens (≈ 2400 chars) for local models
        - overlap: ~120 tokens (≈ 480 chars) to maintain context
        - Sentence-aware splitting to avoid breaking mid-thought
        - Paragraph boundary preferred over sentence boundary
        
        Args:
            text: Input text
            chunk_size: Target tokens (default 600)
            overlap: Overlap tokens (default 120)
        
        Returns:
            List of text chunks
        """
        if not text.strip():
            return []

        # Convert token counts to character estimates
        chunk_chars = chunk_size * FileProcessor.CHARS_PER_TOKEN  # ~2400 for 600 tokens
        overlap_chars = overlap * FileProcessor.CHARS_PER_TOKEN    # ~480 for 120 tokens

        if len(text) <= chunk_chars:
            return [text.strip()]

        chunks: list[str] = []
        sentences = FileProcessor._split_into_sentences(text)
        
        current_chunk = []
        current_length = 0

        for sentence in sentences:
            sentence_length = len(sentence)
            
            # If adding this sentence exceeds chunk size
            if current_length + sentence_length > chunk_chars and current_chunk:
                # Save current chunk
                chunk_text = " ".join(current_chunk).strip()
                if chunk_text:
                    chunks.append(chunk_text)
                
                # Start new chunk with overlap
                # Keep last few sentences for context
                overlap_sentences = []
                overlap_length = 0
                for s in reversed(current_chunk):
                    if overlap_length + len(s) <= overlap_chars:
                        overlap_sentences.insert(0, s)
                        overlap_length += len(s) + 1
                    else:
                        break
                
                current_chunk = overlap_sentences + [sentence]
                current_length = sum(len(s) for s in current_chunk) + len(current_chunk)
            else:
                current_chunk.append(sentence)
                current_length += sentence_length + 1

        # Add final chunk
        if current_chunk:
            chunk_text = " ".join(current_chunk).strip()
            if chunk_text:
                chunks.append(chunk_text)

        return chunks

    @staticmethod
    def _split_into_sentences(text: str) -> list[str]:
        """Split text into sentences, handling common abbreviations."""
        # Simple but effective sentence splitting
        text = text.replace("\n", " ")
        
        # Handle common abbreviations
        text = re.sub(r"Dr\.", "Dr", text)
        text = re.sub(r"Mr\.", "Mr", text)
        text = re.sub(r"Mrs\.", "Mrs", text)
        text = re.sub(r"Ms\.", "Ms", text)
        text = re.sub(r"Prof\.", "Prof", text)
        text = re.sub(r"Jr\.", "Jr", text)
        text = re.sub(r"Sr\.", "Sr", text)
        text = re.sub(r"e\.g\.", "eg", text)
        text = re.sub(r"i\.e\.", "ie", text)
        text = re.sub(r"etc\.", "etc", text)

        # Split on sentence boundaries
        sentences = re.split(r"(?<=[.!?])\s+(?=[A-Z])", text)
        
        # Restore abbreviations and clean
        sentences = [
            s.replace("Dr ", "Dr. ")
            .replace("Mr ", "Mr. ")
            .replace("Mrs ", "Mrs. ")
            .replace("Ms ", "Ms. ")
            .replace("Prof ", "Prof. ")
            .replace("Jr ", "Jr. ")
            .replace("Sr ", "Sr. ")
            .replace("eg ", "e.g. ")
            .replace("ie ", "i.e. ")
            .replace("etc ", "etc. ")
            .strip()
            for s in sentences
        ]
        
        return [s for s in sentences if s]


    @staticmethod
    def validate_file(file_path: str) -> bool:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        if path.suffix.lower() not in FileProcessor.SUPPORTED_FORMATS:
            raise ValueError(f"Unsupported format: {path.suffix}. Supported: {FileProcessor.SUPPORTED_FORMATS}")
        if path.stat().st_size > FileProcessor.MAX_FILE_SIZE:
            raise ValueError(f"File too large: {path.stat().st_size} bytes. Max: {FileProcessor.MAX_FILE_SIZE}")
        return True

    @staticmethod
    def clean_text(text: str) -> str:
        # Collapse excessive whitespace but preserve paragraph breaks
        text = re.sub(r"[ \t]+", " ", text)          # collapse horizontal whitespace
        text = re.sub(r"\n{3,}", "\n\n", text)        # max 2 consecutive newlines
        return text.strip()
